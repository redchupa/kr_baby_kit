"""Convert KDCA growth-chart sources to custom_components/kr_baby_kit/data/growth_chart_kr.json.

Two input modes are supported:

1) CSV (preferred for diffs):
       python scripts/build_growth_chart.py path/to/*.csv

   Required columns: sex, kind, age_months, L, M, S
   sex   ∈ {male, female}
   kind  ∈ {height, weight, head, bmi}

2) XLSX from the public KDCA bundle ("성장도표+데이터+테이블"):
       python scripts/build_growth_chart.py path/to/growth_chart.xlsx --xlsx

   The bundle ships 7 sheets in a fixed order:
       [0] 연령별 신장   → samples.{sex}.height       (cols: 성별|만|개월|L|M|S)
       [1] 연령별 체중   → samples.{sex}.weight       (cols: 성별|만|개월|L|M|S)
       [2] 연령별 체질량지수 → samples.{sex}.bmi      (cols: 성별|만|개월|L|M|S)
       [3] 신장별 체중 (<2y)   → weight_for_length.{sex}.under_2y  (cols: 성별|누운키|L|M|S)
       [4] 신장별 체중 (2-3y)  → weight_for_length.{sex}.age_2_to_3
       [5] 신장별 체중 (3y+)   → weight_for_length.{sex}.age_3_plus
       [6] 연령별 머리둘레 → samples.{sex}.head

   For sheets 0-2 and 6 the script keeps L/M/S keyed by age_months.
   For sheets 3-5 it keeps L/M/S keyed by length_cm (recumbent for <2y per
   KDCA convention, standing for 2y+).

Source: 질병관리청 (KDCA) 2017 소아청소년 성장도표 — 공공누리 제1유형.
The raw XLSX/CSV is NOT bundled in the repo (license requires attribution, not
mirror-redistribution). Download from
    https://www.data.go.kr/data/15076588/fileData.do
or the original KDCA page and pass the path to this script.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = REPO_ROOT / "custom_components" / "kr_baby_kit" / "data" / "growth_chart_kr.json"

_REQUIRED = {"sex", "kind", "age_months", "L", "M", "S"}
_ALLOWED_SEX = {"male", "female"}
_ALLOWED_KIND = {"height", "weight", "head", "bmi"}

# KDCA XLSX bundle: age-indexed sheets (sheet index → kind).
_XLSX_KIND_BY_SHEET = {
    0: "height",
    1: "weight",
    2: "bmi",
    6: "head",
}
# Length-indexed sheets (weight-for-length).
_XLSX_WFL_BAND_BY_SHEET = {
    3: "under_2y",
    4: "age_2_to_3",
    5: "age_3_plus",
}
_XLSX_SEX_MAP = {1: "male", 2: "female"}

# Age-sheet column layout: 성별 | 만(년) | 개월수 | L | M | S | percentile...
_AGE_COL_SEX = 1
_AGE_COL_AGE_MONTHS = 3
_AGE_COL_L = 4
_AGE_COL_M = 5
_AGE_COL_S = 6

# Weight-for-length sheet layout: 성별 | 길이(cm) | L | M | S | percentile...
_WFL_COL_SEX = 1
_WFL_COL_LENGTH = 2
_WFL_COL_L = 3
_WFL_COL_M = 4
_WFL_COL_S = 5

_XLSX_HEADER_ROWS = 2


def _read_csv_rows(paths: list[Path]) -> Iterable[tuple[str, str, dict]]:
    for path in paths:
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            header = {h.strip() for h in (reader.fieldnames or [])}
            missing = _REQUIRED - header
            if missing:
                raise SystemExit(
                    f"{path}: missing columns {sorted(missing)}; got {sorted(header)}"
                )
            for line_no, row in enumerate(reader, start=2):
                sex = row["sex"].strip().lower()
                kind = row["kind"].strip().lower()
                if sex not in _ALLOWED_SEX:
                    raise SystemExit(f"{path}:{line_no}: bad sex {sex!r}")
                if kind not in _ALLOWED_KIND:
                    raise SystemExit(f"{path}:{line_no}: bad kind {kind!r}")
                try:
                    yield sex, kind, {
                        "age_months": float(row["age_months"]),
                        "L": float(row["L"]),
                        "M": float(row["M"]),
                        "S": float(row["S"]),
                    }
                except ValueError as err:
                    raise SystemExit(f"{path}:{line_no}: {err}") from err


def _read_xlsx_age_rows(wb, path: Path) -> Iterable[tuple[str, str, dict]]:
    sheets = wb.sheetnames
    for idx, kind in _XLSX_KIND_BY_SHEET.items():
        if idx >= len(sheets):
            raise SystemExit(
                f"{path}: expected sheet index {idx} for kind={kind} but workbook has {len(sheets)} sheets"
            )
        ws = wb[sheets[idx]]
        for row_no, row in enumerate(
            ws.iter_rows(min_row=_XLSX_HEADER_ROWS + 1, values_only=True),
            start=_XLSX_HEADER_ROWS + 1,
        ):
            if not row or row[_AGE_COL_SEX - 1] in (None, ""):
                continue
            try:
                sex_code = int(row[_AGE_COL_SEX - 1])
            except (TypeError, ValueError):
                continue
            sex = _XLSX_SEX_MAP.get(sex_code)
            if sex is None:
                continue
            age_m = row[_AGE_COL_AGE_MONTHS - 1]
            L = row[_AGE_COL_L - 1]
            M = row[_AGE_COL_M - 1]
            S = row[_AGE_COL_S - 1]
            if any(v is None for v in (age_m, L, M, S)):
                continue
            try:
                yield sex, kind, {
                    "age_months": float(age_m),
                    "L": float(L),
                    "M": float(M),
                    "S": float(S),
                }
            except (TypeError, ValueError) as err:
                raise SystemExit(
                    f"{path}:sheet={idx}:row={row_no}: bad numeric value {err}"
                ) from err


def _read_xlsx_wfl_rows(wb, path: Path) -> Iterable[tuple[str, str, dict]]:
    """Yield (sex, band, {length_cm, L, M, S}) for the weight-for-length sheets."""
    sheets = wb.sheetnames
    for idx, band in _XLSX_WFL_BAND_BY_SHEET.items():
        if idx >= len(sheets):
            continue  # band absent — skip silently
        ws = wb[sheets[idx]]
        for row_no, row in enumerate(
            ws.iter_rows(min_row=_XLSX_HEADER_ROWS + 1, values_only=True),
            start=_XLSX_HEADER_ROWS + 1,
        ):
            if not row or row[_WFL_COL_SEX - 1] in (None, ""):
                continue
            try:
                sex_code = int(row[_WFL_COL_SEX - 1])
            except (TypeError, ValueError):
                continue
            sex = _XLSX_SEX_MAP.get(sex_code)
            if sex is None:
                continue
            length = row[_WFL_COL_LENGTH - 1]
            L = row[_WFL_COL_L - 1]
            M = row[_WFL_COL_M - 1]
            S = row[_WFL_COL_S - 1]
            if any(v is None for v in (length, L, M, S)):
                continue
            try:
                yield sex, band, {
                    "length_cm": float(length),
                    "L": float(L),
                    "M": float(M),
                    "S": float(S),
                }
            except (TypeError, ValueError) as err:
                raise SystemExit(
                    f"{path}:sheet={idx}:row={row_no}: bad numeric value {err}"
                ) from err


def _open_xlsx(path: Path):
    try:
        import openpyxl
    except ImportError as err:
        raise SystemExit(
            "openpyxl is required for --xlsx mode. Install with: pip install openpyxl"
        ) from err
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _build(paths: list[Path], xlsx: bool) -> dict:
    samples: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    wfl: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))

    if xlsx:
        if len(paths) != 1:
            raise SystemExit("--xlsx mode accepts exactly one workbook")
        wb = _open_xlsx(paths[0])
        for sex, kind, row in _read_xlsx_age_rows(wb, paths[0]):
            samples[sex][kind].append(row)
        for sex, band, row in _read_xlsx_wfl_rows(wb, paths[0]):
            wfl[sex][band].append(row)
    else:
        for sex, kind, row in _read_csv_rows(paths):
            samples[sex][kind].append(row)

    for sex in samples:
        for kind in samples[sex]:
            seen: dict[float, dict] = {}
            for r in samples[sex][kind]:
                seen[r["age_months"]] = r
            samples[sex][kind] = sorted(seen.values(), key=lambda r: r["age_months"])

    for sex in wfl:
        for band in wfl[sex]:
            seen_l: dict[float, dict] = {}
            for r in wfl[sex][band]:
                seen_l[r["length_cm"]] = r
            wfl[sex][band] = sorted(seen_l.values(), key=lambda r: r["length_cm"])

    with DATA_FILE.open(encoding="utf-8") as fh:
        out = json.load(fh)
    out["samples"] = {sex: dict(kinds) for sex, kinds in samples.items()}
    if wfl:
        out["weight_for_length"] = {sex: dict(bands) for sex, bands in wfl.items()}
    out["_meta"]["note"] = (
        f"Built from {len(paths)} {'xlsx' if xlsx else 'csv'} file(s) via "
        "scripts/build_growth_chart.py. Source: KDCA 2017 / 공공데이터포털."
    )
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("inputs", nargs="+", type=Path, help="CSV or XLSX file(s).")
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Treat input as a single KDCA XLSX bundle (default: CSV).",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=DATA_FILE,
        help=f"Output JSON path (default: {DATA_FILE}).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print row counts but do not write JSON.",
    )
    args = parser.parse_args(argv)

    built = _build(args.inputs, xlsx=args.xlsx)
    counts = {
        "age_indexed": {
            sex: {kind: len(rows) for kind, rows in kinds.items()}
            for sex, kinds in built["samples"].items()
        },
        "weight_for_length": {
            sex: {band: len(rows) for band, rows in bands.items()}
            for sex, bands in built.get("weight_for_length", {}).items()
        },
    }
    print(json.dumps(counts, indent=2, ensure_ascii=False))

    if args.dry_run:
        return 0

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as fh:
        json.dump(built, fh, ensure_ascii=False, indent=2)
        fh.write("\n")
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
